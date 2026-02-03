from functools import partial
import io
import base64
from typing import List, Dict, Optional, Any
import asyncio
from loguru import logger
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from PIL import Image
from concurrent.futures import ThreadPoolExecutor
from pydantic_core import to_json
import xlrd
from xlrd.book import Book
from langchain_core.documents import Document
from .base import BaseDocumentLoader, ChunkType


class MSExcelLoader(BaseDocumentLoader):
    def __init__(self, include_images=True) -> None:
        self.include_images = include_images
        self.workbook: Book | Workbook | None = None
        self.filename: str | None = None
        self.executor = ThreadPoolExecutor(max_workers=5)

    async def load_excel(self, source):
        try:
            if isinstance(source, bytes):
                source = io.BytesIO(source)
            self.workbook = await asyncio.get_event_loop().run_in_executor(
                self.executor, partial(load_workbook, data_only=False), source
            )
        except (InvalidFileException, BadZipFile):
            try:
                if isinstance(source, str):
                    self.workbook = await asyncio.get_event_loop().run_in_executor(
                        self.executor, xlrd.open_workbook, source
                    )
                elif isinstance(source, (io.BytesIO, bytes)):
                    if isinstance(source, io.BytesIO):
                        source = source.getvalue()
                    self.workbook = await asyncio.get_event_loop().run_in_executor(
                        self.executor,
                        partial(xlrd.open_workbook, file_contents=source),
                    )
            except Exception as e:
                raise ValueError(f"Unable to load Excel file: {str(e)}")

    @property
    def metadata(self) -> Dict[str, Any]:
        if not self.workbook:
            raise ValueError("Excel workbook not loaded. Call load_excel() first.")
        
        base_metadata = {"filename": self.filename}
        
        if hasattr(self.workbook, "properties"):
            # Update with other properties if available, but keep filename from self.filename
            props = {
                "created_at": self.workbook.properties.created,
                "modified": self.workbook.properties.modified,
                "author": self.workbook.properties.creator,
            }
            # Only use workbook title if self.filename is somehow missing
            if not self.filename and self.workbook.properties.title:
                base_metadata["filename"] = self.workbook.properties.title
            
            base_metadata.update(props)
            
        return base_metadata

    def get_sheet_names(self) -> List[str]:
        if not self.workbook:
            raise ValueError("Excel workbook not loaded. Call load_excel() first.")
        return (
            self.workbook.sheetnames
            if hasattr(self.workbook, "sheetnames")
            else self.workbook.sheet_names()
        )

    async def extract_sheet_data(self, sheet_name: str) -> List[Document]:
        if not self.workbook:
            raise ValueError("Excel workbook not loaded. Call load_excel() first.")

        sheet = (
            self.workbook[sheet_name]
            if hasattr(self.workbook, "__getitem__")
            else self.workbook[sheet_name]
        )

        data = {
            "text_content": await self._extract_text_content(sheet),
            "images": await self._extract_images(sheet),
            "charts": await self._extract_charts(sheet),
            "tables": await self._extract_tables(sheet),
            "cell_formats": await self._extract_cell_formats(sheet),
            "merged_cells": await self._extract_merged_cells(sheet),
            "hidden_rows_cols": await self._extract_hidden_rows_cols(sheet),
            "formulas": await self._extract_formulas(sheet),
            "conditional_formatting": await self._extract_conditional_formatting(sheet),
            "data_validation": await self._extract_data_validation(sheet),
            "sheet_properties": await self._extract_sheet_properties(sheet),
        }

        return data

    async def _extract_text_content(self, sheet) -> List[Document]:
        if hasattr(sheet, "iter_rows"):
            df = pd.DataFrame(
                [[cell.value for cell in row] for row in sheet.iter_rows()]
            )
        else:
            df = pd.DataFrame(
                [[cell.value for cell in row] for row in sheet.get_rows()]
            )
        df.dropna(axis=0, how="all", inplace=True)
        df.dropna(axis=1, thresh=int(0.9 * df.shape[0]), inplace=True)
        df.fillna("missing", inplace=True)
        df.columns = df.iloc[0]
        df = df[1:]
        if df.empty:
            return []
        rows = df.apply(lambda x: str(x.to_dict()), axis=1).tolist()
        return [
            Document(
                page_content=to_json(row),
                metadata={
                    "type": ChunkType.text,
                    "sheet_name": sheet.title
                    if hasattr(sheet, "title")
                    else sheet.name,
                    "row_number": row_idx + 1,
                    **self.metadata,
                },
            )
            for row_idx, row in enumerate(rows)
        ]

    async def _extract_images(self, sheet) -> List[Document]:
        images = []
        if hasattr(sheet, "_images"):
            for image in sheet._images:
                img_data = await self.process_image(image)
                images.append(img_data)
        return images

    async def _extract_charts(self, sheet) -> List[Dict[str, Any]]:
        charts = []
        if hasattr(sheet, "_charts"):
            for chart in sheet._charts:
                chart_data = {
                    "type": type(chart).__name__,
                    "title": chart.title,
                    "x_axis": chart.x_axis.title,
                    "y_axis": chart.y_axis.title,
                    "series": [
                        {"name": s.title, "values": s.values} for s in chart.series
                    ],
                }
                charts.append(chart_data)
        return charts

    async def _extract_tables(self, sheet) -> List[Dict[str, Any]]:
        tables = []
        if hasattr(sheet, "tables"):
            for table in sheet.tables.values():
                table_data = {
                    "name": table.name,
                    "range": table.ref,
                    "columns": [
                        {"name": col.name, "total_function": col.totalsRowFunction}
                        for col in table.tableColumns
                    ],
                }
                tables.append(table_data)
        return tables

    async def _extract_cell_formats(self, sheet) -> Dict[str, Dict[str, Any]]:
        formats = {}
        if hasattr(sheet, "iter_rows"):
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.has_style:
                        formats[cell.coordinate] = {
                            "font": self._font_to_dict(cell.font),
                            "fill": self._fill_to_dict(cell.fill),
                            "border": self._border_to_dict(cell.border),
                            "number_format": cell.number_format,
                        }
        return formats

    async def _extract_merged_cells(self, sheet) -> List[str]:
        if hasattr(sheet, "merged_cells"):
            return [str(merged_range) for merged_range in sheet.merged_cells]
        return []

    async def _extract_hidden_rows_cols(self, sheet) -> Dict[str, List[int]]:
        hidden = {"rows": [], "columns": []}
        if hasattr(sheet, "row_dimensions"):
            hidden["rows"] = [
                idx for idx, row_dim in sheet.row_dimensions.items() if row_dim.hidden
            ]
        if hasattr(sheet, "column_dimensions"):
            hidden["columns"] = [
                get_column_letter(idx)
                for idx, col_dim in sheet.column_dimensions.items()
                if col_dim.hidden
            ]
        return hidden

    async def _extract_formulas(self, sheet) -> Dict[str, str]:
        formulas = {}
        if hasattr(sheet, "iter_rows"):
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formulas[cell.coordinate] = cell.value
        return formulas

    async def _extract_conditional_formatting(self, sheet) -> List[Dict[str, Any]]:
        cf_rules = []
        if hasattr(sheet, "conditional_formatting"):
            for cf_range, rules in sheet.conditional_formatting._cf_rules.items():
                for rule in rules:
                    cf_rules.append(
                        {
                            "range": str(cf_range),
                            "type": rule.type,
                            "formula": rule.formula,
                            "dxf": self._dxf_to_dict(rule.dxf) if rule.dxf else None,
                        }
                    )
        return cf_rules

    async def _extract_data_validation(self, sheet) -> Dict[str, Dict[str, Any]]:
        validations = {}
        if hasattr(sheet, "data_validations"):
            for dv in sheet.data_validations.dataValidation:
                validations[dv.sqref] = {
                    "type": dv.type,
                    "operator": dv.operator,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "allow_blank": dv.allow_blank,
                    "error_title": dv.error_title,
                    "error_message": dv.error,
                }
        return validations

    async def process_image(self, image: XLImage) -> Document:
        img = Image.open(image.ref)
        buffered = io.BytesIO()
        img.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode()

        return Document(
            page_content=img_str,
            metadata={
                "size": img.size,
                "format": img.format,
                "filename": img.filename,
                "type": ChunkType.image,
                **self.metadata,
            },
        )

    async def load_async(
        self, source, sheet_names: Optional[List[str]] = None
    ) -> List[Document]:
        logger.info(f"Loading Excel {source}")
        self.filename = self._get_filename(source)
        file = await self.read_source_async(source)
        await self.load_excel(file)

        if sheet_names is None:
            sheet_names = self.get_sheet_names()

        results = []
        for sheet_name in sheet_names:
            sheet = (
                self.workbook[sheet_name]
                if hasattr(self.workbook, "__getitem__")
                else self.workbook[sheet_name]
            )
            if sheet_name in self.get_sheet_names():
                results.extend(await self._extract_text_content(sheet))
                if self.include_images:
                    results.extend(await self._extract_images(sheet))
                    # results.extend(await self._extract_charts(sheet))
            else:
                logger.warning(
                    f"Warning: Sheet '{sheet_name}' not found in the workbook."
                )

        return results

    def _font_to_dict(self, font: Font) -> Dict[str, Any]:
        return {
            "name": font.name,
            "size": font.size,
            "bold": font.bold,
            "italic": font.italic,
            "color": font.color.rgb if font.color else None,
        }

    def _fill_to_dict(self, fill: PatternFill) -> Dict[str, Any]:
        return {
            "fill_type": fill.fill_type,
            "start_color": fill.start_color.rgb if fill.start_color else None,
            "end_color": fill.end_color.rgb if fill.end_color else None,
        }

    def _border_to_dict(self, border: Border) -> Dict[str, Any]:
        return {
            "left": self._side_to_dict(border.left),
            "right": self._side_to_dict(border.right),
            "top": self._side_to_dict(border.top),
            "bottom": self._side_to_dict(border.bottom),
        }

    def _side_to_dict(self, side: Side) -> Dict[str, Any]:
        return {
            "style": side.style,
            "color": side.color.rgb if side.color else None,
        }

    def _dxf_to_dict(self, dxf) -> Dict[str, Any]:
        return {
            "font": self._font_to_dict(dxf.font) if dxf.font else None,
            "fill": self._fill_to_dict(dxf.fill) if dxf.fill else None,
            "border": self._border_to_dict(dxf.border) if dxf.border else None,
        }

    def __del__(self):
        if hasattr(self, "workbook") and self.workbook:
            if hasattr(self.workbook, "close"):
                self.workbook.close()
        if hasattr(self, "executor"):
            self.executor.shutdown()
